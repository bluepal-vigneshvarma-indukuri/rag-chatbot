"""
Extract text and tables from uploaded files.

Tables are rendered as Markdown so the LLM can read structured data clearly.

Supported formats:
  PDF   — text + tables via pdfplumber (falls back to pypdf if unavailable)
  DOCX  — paragraphs + tables in document order via python-docx
  XLSX  — each sheet rendered as a Markdown table via openpyxl
  CSV   — rendered as a Markdown table via built-in csv module
  HTML  — tables converted to Markdown in-place via BeautifulSoup
  Other — plain text (code, JSON, YAML, TXT, MD, etc.)
"""
import io
import csv as _csv
import chardet


def parse_file(filename: str, content: bytes) -> str:
    """Return text (with tables as Markdown) for any supported file type."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "txt"

    if ext == "pdf":
        return _parse_pdf(content)
    elif ext == "docx":
        return _parse_docx(content)
    elif ext == "doc":
        raise ValueError(".doc files are not supported. Please save as .docx.")
    elif ext in ("xlsx", "xls"):
        return _parse_excel(content)
    elif ext == "csv":
        return _parse_csv(content)
    elif ext in ("html", "htm"):
        return _parse_html(content)
    else:
        return _parse_text(content)


# ── PDF ───────────────────────────────────────────────────────────────────────

def _parse_pdf(content: bytes) -> str:
    """Use pdfplumber for table-aware extraction; fall back to pypdf."""
    try:
        return _parse_pdf_pdfplumber(content)
    except Exception:
        return _parse_pdf_fallback(content)


def _parse_pdf_pdfplumber(content: bytes) -> str:
    import pdfplumber

    pages_out = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            parts = []

            # Detect tables on this page
            tables_found = page.find_tables()

            if tables_found:
                # Collect bounding boxes of all table cells so we can
                # exclude that area from plain-text extraction.
                table_bboxes = [t.bbox for t in tables_found]

                # Filter page chars to those NOT inside any table bbox
                def _outside_tables(obj):
                    if obj.get("object_type") != "char":
                        return True
                    x0, top, x1, bottom = (
                        obj.get("x0", 0), obj.get("top", 0),
                        obj.get("x1", 0), obj.get("bottom", 0),
                    )
                    for bx0, btop, bx1, bbottom in table_bboxes:
                        if x0 >= bx0 - 2 and x1 <= bx1 + 2 and top >= btop - 2 and bottom <= bbottom + 2:
                            return False
                    return True

                outside_page = page.filter(_outside_tables)
                plain_text = outside_page.extract_text() or ""
            else:
                plain_text = page.extract_text() or ""

            if plain_text.strip():
                parts.append(plain_text.strip())

            # Append each table as Markdown (after the surrounding text)
            for tbl_obj in tables_found:
                rows = tbl_obj.extract()
                md = _table_to_markdown(rows)
                if md:
                    parts.append(md)

            if parts:
                pages_out.append("\n\n".join(parts))

    return "\n\n".join(pages_out)


def _parse_pdf_fallback(content: bytes) -> str:
    """Plain text extraction using pypdf (no table awareness)."""
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(content))
    pages = []
    for page in reader.pages:
        text = page.extract_text() or ""
        if text.strip():
            pages.append(text.strip())
    return "\n\n".join(pages)


# ── DOCX ──────────────────────────────────────────────────────────────────────

def _parse_docx(content: bytes) -> str:
    """
    Iterate the document body in order so tables stay in context with
    surrounding paragraphs. Tables are rendered as Markdown.
    """
    from docx import Document
    from docx.oxml.ns import qn
    from docx.text.paragraph import Paragraph
    from docx.table import Table

    doc = Document(io.BytesIO(content))
    parts = []

    for child in doc.element.body:
        tag = child.tag

        if tag == qn("w:p"):
            para = Paragraph(child, doc)
            text = para.text.strip()
            if text:
                parts.append(text)

        elif tag == qn("w:tbl"):
            table = Table(child, doc)
            rows = []
            for row in table.rows:
                # Merge duplicate cells that span columns (python-docx repeats them)
                seen, cells = set(), []
                for cell in row.cells:
                    cid = id(cell._tc)
                    if cid not in seen:
                        seen.add(cid)
                        cells.append(cell.text.strip().replace("\n", " "))
                rows.append(cells)
            md = _table_to_markdown(rows)
            if md:
                parts.append(md)

    return "\n\n".join(parts)


# ── XLSX / XLS ────────────────────────────────────────────────────────────────

def _parse_excel(content: bytes) -> str:
    """Render every sheet as a Markdown table."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            # Skip completely empty rows
            if any(cell is not None and str(cell).strip() for cell in row):
                cells = [str(cell).strip() if cell is not None else "" for cell in row]
                rows.append(cells)

        if rows:
            md = _table_to_markdown(rows)
            parts.append(f"## Sheet: {sheet_name}\n\n{md}")

    wb.close()
    return "\n\n".join(parts)


# ── CSV ───────────────────────────────────────────────────────────────────────

def _parse_csv(content: bytes) -> str:
    """Render CSV as a Markdown table."""
    decoded = _decode_bytes(content)
    reader = _csv.reader(io.StringIO(decoded))
    rows = []
    for row in reader:
        if any(cell.strip() for cell in row):
            rows.append([cell.strip() for cell in row])
    return _table_to_markdown(rows) if rows else ""


# ── HTML ──────────────────────────────────────────────────────────────────────

def _parse_html(content: bytes) -> str:
    """
    Replace <table> elements with Markdown in-place, then extract all text.
    This keeps the table in context relative to surrounding headings/paragraphs.
    """
    from bs4 import BeautifulSoup

    decoded = _decode_bytes(content)
    soup = BeautifulSoup(decoded, "html.parser")

    # Drop non-content tags
    for tag in soup(["script", "style", "head", "nav", "footer"]):
        tag.decompose()

    # Convert every <table> to its Markdown equivalent in-place
    for table_tag in soup.find_all("table"):
        rows = []
        for tr in table_tag.find_all("tr"):
            cells = [
                td.get_text(separator=" ", strip=True)
                for td in tr.find_all(["td", "th"])
            ]
            if cells:
                rows.append(cells)
        md = _table_to_markdown(rows)
        if md:
            table_tag.replace_with(f"\n\n{md}\n\n")
        else:
            table_tag.decompose()

    text = soup.get_text(separator="\n", strip=True)
    # Collapse excessive blank lines
    import re
    return re.sub(r"\n{3,}", "\n\n", text).strip()


# ── Plain text ────────────────────────────────────────────────────────────────

def _parse_text(content: bytes) -> str:
    """Decode arbitrary text/code files (TXT, MD, JSON, Python, etc.)."""
    return _decode_bytes(content)


def _decode_bytes(content: bytes) -> str:
    detected = chardet.detect(content)
    encoding = detected.get("encoding") or "utf-8"
    try:
        return content.decode(encoding)
    except (UnicodeDecodeError, LookupError):
        return content.decode("utf-8", errors="replace")


# ── Table → Markdown helper ───────────────────────────────────────────────────

def _table_to_markdown(table: list) -> str:
    """
    Convert a list-of-lists table into a Markdown table string.
    - First row is treated as the header.
    - None cells become empty strings.
    - Newlines inside cells are collapsed to spaces.
    """
    if not table:
        return ""

    # Normalise cells
    rows = []
    for raw_row in table:
        if raw_row is None:
            continue
        cells = [
            str(cell if cell is not None else "").strip().replace("\n", " ").replace("|", "\\|")
            for cell in raw_row
        ]
        rows.append(cells)

    if not rows:
        return ""

    # Make all rows the same width
    max_cols = max(len(r) for r in rows)
    rows = [r + [""] * (max_cols - len(r)) for r in rows]

    # Column widths (minimum 3 so the separator dashes are valid Markdown)
    col_widths = [
        max(3, max(len(r[i]) for r in rows))
        for i in range(max_cols)
    ]

    def _row_line(row):
        padded = [row[i].ljust(col_widths[i]) for i in range(max_cols)]
        return "| " + " | ".join(padded) + " |"

    lines = [_row_line(rows[0])]
    lines.append("| " + " | ".join("-" * w for w in col_widths) + " |")
    for row in rows[1:]:
        lines.append(_row_line(row))

    return "\n".join(lines)
